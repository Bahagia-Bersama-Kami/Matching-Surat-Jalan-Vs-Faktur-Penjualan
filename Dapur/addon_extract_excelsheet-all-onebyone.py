import openpyxl
import os
import sys

def list_excel_files():
    return [f for f in os.listdir('.') if f.endswith(('.xlsx', '.xlsm'))]

def main():
    print("=== EXCEL SHEET EXTRACTOR (FULL FORMAT) ===")
    
    files = list_excel_files()
    if not files:
        print("Tidak ditemukan file Excel (.xlsx/.xlsm) di folder ini.")
        input("Tekan Enter untuk keluar...")
        return

    print("\nFile yang ditemukan:")
    for i, f in enumerate(files, 1):
        print(f"{i}. {f}")

    try:
        selection = input("\nPilih nomor file: ")
        file_idx = int(selection) - 1
        selected_file = files[file_idx]
    except (ValueError, IndexError):
        print("Pilihan tidak valid.")
        input("Tekan Enter untuk keluar...")
        return
        
    print(f"\nMenganalisis file: {selected_file}...")
    
    try:
        temp_wb = openpyxl.load_workbook(selected_file, read_only=True, data_only=True)
        all_sheets = temp_wb.sheetnames
        temp_wb.close()
    except Exception as e:
        print(f"Gagal membaca file: {e}")
        input("Tekan Enter untuk keluar...")
        return

    print(f"Ditemukan {len(all_sheets)} sheet. Memulai ekstraksi...")

    for target_sheet in all_sheets:
        print(f"--> Sedang memproses sheet: {target_sheet}")
        
        try:
            wb_process = openpyxl.load_workbook(selected_file)
            
            for s_name in wb_process.sheetnames:
                if s_name != target_sheet:
                    sheet_to_remove = wb_process[s_name]
                    wb_process.remove(sheet_to_remove)
                    
            clean_name = "".join([c for c in target_sheet if c.isalnum() or c in (' ', '-', '_')]).strip()
            if not clean_name:
                clean_name = "Unnamed_Sheet"
                
            output_name = f"{clean_name}.xlsx"
            
            wb_process.save(output_name)
            
        except Exception as e:
            print(f"    Gagal mengekstrak {target_sheet}: {e}")

    print("\n--> SELESAI! Semua sheet telah diekstrak.")
    input("Tekan Enter untuk keluar...")

if __name__ == "__main__":
    main()