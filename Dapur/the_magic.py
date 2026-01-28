import pandas as pd
import os
import warnings
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

def get_month_index(sheet_name):
    months = [
        'jan', 'feb', 'mar', 'apr', 'mei', 'jun', 
        'jul', 'ags', 'aug', 'sep', 'okt', 'nov', 'des'
    ]
    name_lower = sheet_name.lower()
    for i, m in enumerate(months):
        if m in name_lower:
            return i
    return 999

def format_indo(number):
    try:
        if pd.isna(number):
            return "0,00"
        s = "{:,.2f}".format(float(number))
        table = s.maketrans({',': '.', '.': ','})
        return s.translate(table)
    except:
        return str(number)

def style_range(ws, cell_range, color_hex):
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    font = Font(bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

def auto_fit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

def proses_data_excel():
    nama_file_input = 'HASIL_EKSTRAK_GABUNGAN.xlsx'
    nama_file_output = 'HASIL_ANALISIS_GABUNGAN.xlsx'
    
    if not os.path.exists(nama_file_input):
        print("--> File GABUNGAN.xlsx tidak ditemukan.")
        return

    print(f"--> Membaca file: {nama_file_input}")
    
    try:
        all_sheets = pd.read_excel(nama_file_input, sheet_name=None)
    except Exception as e:
        print(f"--> Gagal membaca file: {e}")
        return

    sorted_sheet_names = sorted(all_sheets.keys(), key=get_month_index)
    
    writer = pd.ExcelWriter(nama_file_output, engine='openpyxl')
    
    for nama_sheet in sorted_sheet_names:
        df = all_sheets[nama_sheet]
        print(f"--> Memproses Sheet: {nama_sheet}")
        
        df.columns = [str(c).strip() for c in df.columns]
        
        if 'Keterangan' not in df.columns or 'Nomor Faktur' not in df.columns:
            print(f"--> Sheet {nama_sheet} dilewati karena kolom tidak lengkap.")
            continue

        df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
        df['Kredit'] = pd.to_numeric(df['Kredit'], errors='coerce').fillna(0)
        df['Tanggal'] = pd.to_datetime(df['Tanggal'], dayfirst=True, errors='coerce')
        
        total_debit = df[df['Keterangan'] == 'Pengiriman Pesanan']['Debit'].sum()
        total_kredit = df[df['Keterangan'] == 'Faktur Penjualan']['Kredit'].sum()
        selisih_total = total_debit - total_kredit

        data_hasil_a = [
            {'Ringkasan': 'Total Debit (Pengiriman Pesanan)', 'Nilai': format_indo(total_debit)},
            {'Ringkasan': 'Total Kredit (Faktur Penjualan)', 'Nilai': format_indo(total_kredit)},
            {'Ringkasan': 'Selisih', 'Nilai': format_indo(selisih_total)}
        ]
        
        unique_faktur = df['Nomor Faktur'].dropna().unique()
        
        data_hasil_c = []
        data_hasil_b = []

        for faktur in unique_faktur:
            subset = df[df['Nomor Faktur'] == faktur]
            
            row_kirim = subset[subset['Keterangan'] == 'Pengiriman Pesanan']
            row_faktur = subset[subset['Keterangan'] == 'Faktur Penjualan']
            
            val_debit = 0
            val_kredit = 0
            date_kirim = None
            date_faktur = None
            
            has_kirim = not row_kirim.empty
            has_faktur = not row_faktur.empty
            
            if has_kirim:
                val_debit = row_kirim.iloc[0]['Debit']
                date_kirim = row_kirim.iloc[0]['Tanggal']
            
            if has_faktur:
                val_kredit = row_faktur.iloc[0]['Kredit']
                date_faktur = row_faktur.iloc[0]['Tanggal']
            
            selisih = val_debit - val_kredit
            status = "Match"
            
            if not has_kirim or not has_faktur:
                status = "Tidak Lengkap (Pasangan Hilang)"
            elif abs(selisih) > 5:
                status = "Selisih Angka > 5"
            elif date_kirim != date_faktur:
                status = "Tanggal Berbeda"

            if has_kirim:
                r_k = row_kirim.iloc[0].to_dict()
                r_k['Analisis_Status'] = status
                r_k['Analisis_Selisih'] = selisih
                data_hasil_c.append(r_k)
                if status != "Match":
                    data_hasil_b.append(r_k)

            if has_faktur:
                r_f = row_faktur.iloc[0].to_dict()
                r_f['Analisis_Status'] = status
                r_f['Analisis_Selisih'] = selisih
                data_hasil_c.append(r_f)
                if status != "Match":
                    data_hasil_b.append(r_f)

        df_a = pd.DataFrame(data_hasil_a)
        df_b = pd.DataFrame(data_hasil_b)
        df_c = pd.DataFrame(data_hasil_c)
        
        numeric_cols = ['Debit', 'Kredit', 'Analisis_Selisih']
        
        if not df_b.empty:
            if 'Tanggal' in df_b.columns:
                df_b['Tanggal'] = df_b['Tanggal'].dt.strftime('%d/%m/%Y')
            for col in numeric_cols:
                if col in df_b.columns:
                    df_b[col] = df_b[col].apply(format_indo)
            
        if not df_c.empty:
            if 'Tanggal' in df_c.columns:
                df_c['Tanggal'] = df_c['Tanggal'].dt.strftime('%d/%m/%Y')
            for col in numeric_cols:
                if col in df_c.columns:
                    df_c[col] = df_c[col].apply(format_indo)

        current_row = 0
        
        df_a.to_excel(writer, sheet_name=nama_sheet, startrow=current_row + 1, index=False)
        ws = writer.sheets[nama_sheet]
        ws[f'A{current_row + 1}'] = "A. Ringkasan Total"
        ws[f'A{current_row + 1}'].font = Font(bold=True, size=12)
        
        header_row = current_row + 2
        last_col_letter = get_column_letter(len(df_a.columns))
        style_range(ws, f"A{header_row}:{last_col_letter}{header_row}", "90EE90") 
        
        current_row += len(df_a) + 4 
        
        df_b.to_excel(writer, sheet_name=nama_sheet, startrow=current_row + 1, index=False)
        ws[f'A{current_row + 1}'] = "B. Hasil yang Tidak Match"
        ws[f'A{current_row + 1}'].font = Font(bold=True, size=12)
        
        if not df_b.empty:
            header_row = current_row + 2
            last_col_letter = get_column_letter(len(df_b.columns))
            style_range(ws, f"A{header_row}:{last_col_letter}{header_row}", "FFCCCB")
            current_row += len(df_b) + 4
        else:
            ws[f'A{current_row + 3}'] = "(Tidak ada data error)"
            current_row += 5

        df_c.to_excel(writer, sheet_name=nama_sheet, startrow=current_row + 1, index=False)
        ws[f'A{current_row + 1}'] = "C. Data Asli Diurutkan"
        ws[f'A{current_row + 1}'].font = Font(bold=True, size=12)
        
        if not df_c.empty:
            header_row = current_row + 2
            last_col_letter = get_column_letter(len(df_c.columns))
            style_range(ws, f"A{header_row}:{last_col_letter}{header_row}", "ADD8E6")

        auto_fit_columns(ws)

    try:
        writer.close()
        print(f"--> Selesai. Hasil disimpan di: {nama_file_output}")
    except Exception as e:
        print(f"--> Gagal menyimpan file: {e}")

if __name__ == "__main__":
    proses_data_excel()